import openpyxl
import matplotlib.pyplot as plt
from datetime import datetime
from os import path

def plot_short_selling(stock_number):
    today = datetime.today()

    # 讀取 Excel 檔案
    wb = openpyxl.load_workbook(f"C:/temp/stock-log/{stock_number}_{today.strftime('%Y-%m')}.xlsx")
    ws = wb.active

    # 讀取資料（假設第一列是標題，數據從第二列開始）
    date_time = []
    data_series = [[] for _ in range(1)]  # 假設有 1 個數據欄位，在陣列中創造1個空陣列

    # Excel的每一列數據會存在ws.iter_rows，將時間拿出來存在date_time陣列，其他資料放在data_series
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Convert y-m-d H:M:S to m-d
        date_obj = datetime.strptime(str(row[0]), "%Y-%m-%d %H:%M:%S")
        month_day = date_obj.strftime("%m-%d")

        date_time.append(month_day)
        data_series[0].append(row[4])  # 第四攔數據是借券賣出餘額

    # 繪製圖表
    plt.figure(figsize=(10, 5))
    for i, data in enumerate(data_series):
        plt.plot(date_time, data, label=f"balance")

    plt.xlabel("date")
    plt.ylabel("shares")
    plt.title(f"{stock_number} - Short selling")
    plt.legend()
    plt.xticks(rotation=45)
    plt.grid()

    # 儲存為 JPG
    data_dir = "C:/temp/stock-log"
    file_name = f"{stock_number}_{today.strftime('%Y-%m')}.jpg"
    save_path = path.join(data_dir, file_name)
    plt.savefig(save_path, dpi=300, bbox_inches="tight")