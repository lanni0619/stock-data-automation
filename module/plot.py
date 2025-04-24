# Build-in module
from datetime import datetime
from os import path
from threading import Lock
from typing import cast, Generator

# 3rd-party package
import openpyxl
import matplotlib # type: ignore
import matplotlib.pyplot as plt # type: ignore
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

matplotlib.use("Agg")

plot_lock = Lock()

def plot_short_selling(stock_number: str) -> None:
    with plot_lock:
        today:datetime = datetime.today()

        # 讀取 Excel 檔案
        wb:Workbook = openpyxl.load_workbook(f"C:/temp/stock-log/{stock_number}_{today.strftime('%Y-%m')}.xlsx")
        ws:Worksheet = cast(Worksheet, wb.active)

        # 讀取資料（假設第一列是標題，數據從第二列開始）
        date_time:list = []
        data_series:list[list] = [[] for _ in range(2)]  # 假設有 2 個數據欄位，在陣列中創造 2 個空陣列

        # Excel的每一列數據會存在ws.iter_rows，將時間拿出來存在date_time陣列，其他資料放在data_series
        ws_rows:Generator = ws.iter_rows(min_row=2, values_only=True)

        for row in ws_rows:
            # Convert y-m-d H:M:S to m-d
            date_obj = datetime.strptime(str(row[0]), "%Y-%m-%d %H:%M:%S")
            month_day = date_obj.strftime("%m-%d")

            date_time.append(month_day)
            data_series[0].append(round((row[4])))  # 第四欄數據是借券賣出餘額
            data_series[1].append(row[5])  # 第五欄數據是股價

        # 繪製圖表
        plt.figure(figsize=(10, 5))
        for i, data in enumerate(data_series):
            label = "short selling (unit - 1k lot)" if i == 0 else "price (unit - NTD)"
            plt.plot(date_time, data, label=label)

        plt.xlabel("date")
        plt.ylabel("shares")
        plt.title(f"{stock_number} - Relationship between price & short selling")
        plt.legend()
        plt.xticks(rotation=45)
        plt.grid()

        # 儲存為 JPG
        data_dir = "C:/temp/stock-log"
        file_name = f"{stock_number}_{today.strftime('%Y-%m')}.jpg"
        save_path = path.join(data_dir, file_name)
        plt.savefig(save_path, dpi=300, bbox_inches="tight")
        plt.close()