# 3rd-party
import os

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# standard
from datetime import datetime
from os import path
from typing import Optional, cast, Generator, Tuple
import traceback

# self-define
from module.logger import logger
import module.utils as utils
from config.config import config

class ExcelHandler:
    # 0=name_zh_tw 1=stock_code, 2=YY-MM
    PATH_NAME:str = config.get("excel_settings")["root_path"]
    FILE_NAME:str = config.get("excel_settings")["file_name"]
    INSTANCE_CACHE:dict[str, Optional["ExcelHandler"]] = {}
    HEADER_COLS:list[str] = ["created_at", "balance_yest", "selling_today", "return_today", "balance_today", "price"]

    def __init__(self, wb:Workbook, sheet:Worksheet, file_path:str):
        self.wb:Workbook = wb
        self.sheet:Worksheet = sheet
        self.file_path:str = file_path
        self._initialize_sheet()

    @classmethod
    @utils.tic_tok
    @utils.handle_errors
    def create_file(cls, info:dict) -> "ExcelHandler":
        print(info)
        yy_mm = datetime.today().strftime('%Y-%m')
        path_name:str = cls.PATH_NAME.format(info.get("stock_code"))
        file_name:str = cls.FILE_NAME.format(info.get("stock_code"), yy_mm)
        file_path:str = path.join(path_name, file_name)

        if not path.exists(path_name):
            os.makedirs(path_name)

        excel_handler = cls.INSTANCE_CACHE.get(file_path)

        if excel_handler:
            logger.info("[create_file] ExcelHandler already exist!")
            return excel_handler
        else:
            logger.info("[create_file] Create new ExcelHandler")
            wb:Workbook = openpyxl.load_workbook(file_path, data_only=True) if path.exists(file_path) else Workbook()
            sheet:Worksheet = cast(Worksheet, wb.active)

            excel_handler = cls(wb, sheet, file_path)
            cls.INSTANCE_CACHE[file_path] = excel_handler

            return excel_handler

    def _initialize_sheet(self) -> None:
        if not path.exists(self.file_path):
            logger.info("[_initialize_sheet] File not exists, creating ...")
            self.sheet.append(self.__class__.HEADER_COLS)
            self.wb.save(self.file_path)

    def _read_last_row_date(self) -> Optional[str]:
        last_row: int = self.sheet.max_row
        if last_row != 1:
            last_row_date:str = cast(str, self.sheet.cell(last_row, 1).value)[0:10]
            return last_row_date
        else:
            return None

    def _is_duplicate(self) -> bool:
        last_row_date:Optional[str] = self._read_last_row_date()
        today_date:str = datetime.today().strftime("%Y-%m-%d")

        if last_row_date == today_date:
            return True
        return False

    @utils.tic_tok
    @utils.handle_errors
    def save_file(self, stock_dict:dict) -> None:
        if self._is_duplicate():
            logger.warning("[save_file] Duplicate record, stop saving file")
            return
        else:
            new_row:list[str] = [stock_dict['update_time'], stock_dict['balance_yest'], stock_dict['selling_today'], stock_dict['return_today'], stock_dict['balance_today'], stock_dict['price']]
            self.sheet.append(new_row)
            self.wb.save(self.file_path)

    @utils.tic_tok
    @utils.handle_errors
    def read_all_records(self) -> Tuple[list[str], list[list]]:
        wb:Workbook = openpyxl.load_workbook(self.file_path)
        sheet:Worksheet = cast(Worksheet, wb.active)

        # Provide Plot class x, y axis
        data_x:list[str] = []
        data_y:list[list] = [[] for _ in range(2)] # [[], []]

        # Access excel rows start from min_row
        all_rows:Generator = sheet.iter_rows(min_row=2, values_only=True)

        # row[0]=datetime, row[4]=short_selling, row[5]=price
        # assign row[0] to data_x, row[4] to data_y[0], row[5] to data_y[1]
        for row in all_rows:
            # Convert (y-m-d H:M:S) to (m-d)
            x_m_d:str = str(row[0])[0:10]
            data_x.append(x_m_d)
            data_y[0].append(float(row[4])/1000/1000)
            data_y[1].append(row[5])

        return data_x, data_y

if __name__ == "__main__":
    try:
        # 1) Testing duplicate file
        test_info = {"stock_code":2317}
        excel_2317 = ExcelHandler.create_file(test_info)
        excel_2330 = ExcelHandler.create_file(test_info)
        print(excel_2317 == excel_2330) # True
        print(excel_2317.PATH_NAME)

        # 2) Testing _initialize_sheet() - Create a non-exist file
        # excel_2454 = ExcelHandler.create_file(2454)

        # 3) Testing read_all_records()
        # x_2317, y_2317 = excel_2317.read_all_records()
        # print(x_2317)
        # print(y_2317)

    except Exception as e:
        traceback.print_exc()
        print(e)