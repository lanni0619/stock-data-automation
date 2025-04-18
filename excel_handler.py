# 3rd-party
from operator import truediv

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# standard
from datetime import datetime
from os import path
from typing import Optional, cast
import traceback

# self-define
import utils
from logger import logger

class ExcelHandler:
    # 0=stock_code, 1=YY-MM
    FILE_PATH = path.join("C:/temp/stock-log", "{0}_{1}.xlsx")
    INSTANCE_CACHE:dict = {}
    HEADER_COLS:list[str] = ["created_at", "balance_yest", "selling_today", "return_today", "balance_today", "price"]

    def __init__(self, wb:Workbook, sheet:Worksheet, file_path:str):
        self.wb:Workbook = wb
        self.sheet:Worksheet = sheet
        self.file_path:str = file_path
        self._initialize_sheet()

    @classmethod
    @utils.tic_tok
    @utils.handle_errors
    def create_file(cls, stock_code:str) -> Optional[object]:
        yy_mm = datetime.today().strftime('%Y-%m')
        file_path = cls.FILE_PATH.format(stock_code, yy_mm)

        if not cls.INSTANCE_CACHE.get(file_path, None):
            wb:Workbook = openpyxl.load_workbook(file_path, data_only=True) if path.exists(file_path) else Workbook()
            sheet:Worksheet = cast(Worksheet, wb.active)
            cls.INSTANCE_CACHE[file_path] = True
            return cls(wb, sheet, file_path)

        return None

    def _initialize_sheet(self) -> None:
        if not path.exists(self.file_path):
            logger.info("[_initialize_sheet] File not exists, creating ...")
            self.sheet.append(self.__class__.HEADER_COLS)
            self.wb.save(self.file_path)

    def _read_last_row_date(self) -> Optional[str]:
        last_row: int = self.sheet.max_row
        if last_row != 1:
            last_row_date:str = cast(str, self.sheet.cell(last_row, 1).value)[0:10]
            return last_row_date
        else:
            return None

    def _is_duplicate(self) -> bool:
        last_row_date:Optional[str] = self._read_last_row_date()
        today_date:str = datetime.today().strftime("%Y-%m-%d")

        if last_row_date == today_date:
            return True
        return False

    @utils.tic_tok
    @utils.handle_errors
    def save_file(self, stock_dict:dict) -> None:
        if self._is_duplicate():
            logger.error("[save_file] Duplicate record, stop saving file")
            return
        else:
            new_row:list[str] = [stock_dict['update_time'], stock_dict['balance_yest'], stock_dict['selling_today'], stock_dict['return_today'], stock_dict['balance_today'], stock_dict['price']]
            self.sheet.append(new_row)
            self.wb.save(self.file_path)

if __name__ == "__main__":
    try:
        # 1) Testing duplicate file
        excel_2317 = ExcelHandler.create_file(2317)
        excel_2330 = ExcelHandler.create_file(2317)
        print(excel_2317, excel_2330) # object, None

        # 2) Testing _initialize_sheet() - Create a non-exist file
        # excel_2454 = ExcelHandler.create_file(2454)

    except Exception as e:
        traceback.print_exc()
        print(e)

# def save_to_excel(self) -> None:
#     try:
#         logger.info(f"save_to_excel - stock_number = {self._stock_code}")

#         # 1) Preliminary
#         today:datetime = datetime.today()
#         root_path:str = "C:/temp/stock-log"
#         filename:str = os.path.join(root_path, f"{self._stock_code}_{today.strftime('%Y-%m')}.xlsx")
#         attrs:list[str] = ["created_at", "balance_yest", "selling_today", "return_today", "balance_today", "price"]

#         wb = openpyxl.load_workbook(filename, data_only=True) if os.path.exists(filename) else Workbook()
#         sheet:Worksheet = cast(Worksheet, wb.active)

#         # 2) Check if data exist ?
#         if os.path.exists(filename):
#             logger.info(f"save_to_excel - file exists")
#             # Get max_row number which is the index of latest record
#             max_row:int = sheet.max_row
#             if max_row != 1:
#                 # Get last record datetime (string, yyyy-mm-dd)
#                 # cast(type, val) => Tell mypy the type of val
#                 lr_date:str = cast(str, sheet.cell(max_row, 1).value)[0:10]

#                 # 3) Check if data duplicate ?
#                 if lr_date == today.strftime("%Y-%m-%d"):  # compare with today
#                     logger.info("save_to_excel - Duplicate date of record, stop saving data ")
#                     return

#         # 4) if file not exists
#         else:
#             logger.info(f"save_to_excel - file not exists, creating ...")
#             sheet.append(attrs)
#             wb.save(filename)

#         # 5) build new row
#         logger.info(f"save_to_excel - Building new row ...")
#         row_arr:list[Union[str, float]] = []
#         for attr in attrs:
#             value:str = getattr(self, "_"+attr)
#             if attr != "created_at" and value is not None:
#                 row_arr.append(float(value.replace(",", "")))
#             elif attr == "created_at":
#                 row_arr.append(value)
#             else:
#                 row_arr.append("")

#         sheet.append(row_arr)
#         wb.save(filename)

#         logger.info("save_to_excel - finish")

#     except Exception as e:
#         print(e)