# 3rd party package
import requests
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from requests import Response
from bs4 import BeautifulSoup, Tag, NavigableString, ResultSet
# apscheduler don't have official stub
from apscheduler.schedulers.background import BackgroundScheduler # type: ignore
import openpyxl

# Standard module
import json
from datetime import datetime
import os
import time
from typing import Any, Callable, Optional, Union, cast

# Custom module
from module.plot import plot_short_selling
from logger import logger

def validate_stock_code(stock_code:int):
    if not isinstance(stock_code, int) or not (1101 <= stock_code <= 9958):
        raise ValueError(f"Invalid stock_code: {stock_code}. It must be a string of digits between 1101 and 9958.")

def tic_tok(func: Callable) -> Callable:
    # position & keyword argument. The prefix star is meaning Any number.
    def wrapper(*arg, **kwargs) -> Any:
        t1: float = time.time()
        result:Any = func(*arg, **kwargs)
        t2: float = time.time() - t1
        logger.info(f"{func.__name__} took {round(t2, 3)} seconds")
        return result
    return wrapper

def class_to_json(obj: object) -> str:
    #  https://stackoverflow.com/questions/7408647/convert-dynamic-python-object-to-json
    return json.dumps(
        obj,
        default=lambda o: o.__dict__,
        sort_keys=False,
        indent=4
    )

class Stock:
    def __init__(self, stock_code:int, created_at=None, balance_yest=None,
                 selling_today=None, return_today=None, balance_today=None, price=None):

        validate_stock_code(stock_code)
        
        self._stock_code:str = str(stock_code)
        self._created_at:str = created_at
        self._balance_yest:str = balance_yest
        self._selling_today:str = selling_today
        self._return_today:str = return_today
        self._balance_today:str = balance_today
        self._price:str = price

    @property
    def stock_code(self) -> str:
        return self._stock_code

    @stock_code.setter
    def stock_code(self, value:int) -> None:
        if value != self.stock_code:
            logger.error(f"Class stock{self._stock_code} - Can not modify the stock_code property")
        else:
            self._stock_code = str(value)

    @stock_code.deleter
    def stock_code(self):
        # del self._stock_code
        logger.error(f"Class stock{self._stock_code} - Can not delete the stock_code property")

    @tic_tok
    def crawl_price(self) -> Optional[object]:
        try:
            logger.info(f"crawl_stock_price init - stock_number = {self._stock_code}")
            url:str = f"https://tw.stock.yahoo.com/quote/{self._stock_code}.TW"
            web:Response = requests.get(url)
            soup:BeautifulSoup = BeautifulSoup(web.text, "html5lib")

            target_classes: list[str] = [
                "Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-up)",  # 漲
                "Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-down)",  # 跌
                "Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c)",  # 平盤
                "Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-down)",  # 跌停
                "Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-up)",  # 漲停
            ]

            span_tag:Union[Tag, NavigableString, None] = soup.find('span', class_=lambda x: x and any(cls in x for cls in target_classes))
            if isinstance(span_tag, Tag):
                price: str = span_tag.get_text()
                logger.info(f"crawl_stock_price - Get price = {price}")
                self._price = price
                self._created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                return self
            else:
                raise Exception(f"Fail to get {self._stock_code} price")

        except requests.exceptions.RequestException as e:
            logger.error(f"crawl_stock_price - Network error: {e}")
        except AttributeError as e:
            logger.error(f"crawl_stock_price - Parsing error: {e}")
        except Exception as e:
            logger.error(f"crawl_stock_price - unexpected error: {e}")
        return None

    @tic_tok
    def crawl_short_selling(self) -> object:
        try:
            logger.info(f"crawl_short_selling init - stock_number = {self._stock_code}")
            url:str = 'https://www.twse.com.tw/rwd/zh/marginTrading/TWT93U?response=html'
            web:Response = requests.get(url)
            soup:BeautifulSoup = BeautifulSoup(web.text, "html5lib")
            trs:ResultSet = soup.find_all('tr', attrs={"align":"center", "style":"font-size:14px;"})

            for tr in trs:
                if tr.find('td').get_text() == str(self._stock_code):
                    target_info = tr.find_all('td')
                    self._created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    self._balance_yest = target_info[8].get_text()  # 前日餘額
                    self._selling_today = target_info[9].get_text() # 當日賣出
                    self._return_today = target_info[10].get_text()  # 當日還券
                    self._balance_today = target_info[12].get_text() # 今日餘額 
                    logger.info(f"crawl_short_selling - Get info = {self.__dict__}")
                    return self

            raise Exception(f"Could not find {self._stock_code} short_selling info")

        except requests.exceptions.RequestException as e:
            logger.error(f"crawl_short_selling - Network error: {e}")
        except AttributeError as e:
            logger.error(f"crawl_short_selling - Parsing error: {e}")
        except Exception as e:
            logger.error(f"crawl_short_selling - unexpected error: {e}")
        return None

    @tic_tok
    def send_json(self) -> None:
        try:
            logger.info(f"send_json - init - stock_number = {self._stock_code}")
            # Preliminary
            if not all(getattr(self, attr) is not None for attr in vars(self)):
                logger.error("send_json - data incomplete")
                return

            info_json:str = class_to_json(self)
            url:str = "https://discord.com/api/webhooks/1356484738029719573/9GNCPHfl7gcz9BpkkO1xYEYqZ9_D2tWd0dx5sZqx3RTN3HgLFLql47TEgWYEsz0Q4x8g"   
            headers:dict = {"Content-Type": "application/json"}
            data:dict = {"content": info_json, "username": "newmanBot"}

            # Send data to discord
            res:Response = requests.post(url, headers = headers, json = data)

            # Read the response
            if res.status_code in (200, 204):
                logger.info(f"send_json - success")
            else:
                logger.error(f"send_json - fail")

        except requests.exceptions.RequestException as e:
            logger.error(f"send_json - Network error: {e}")
        except AttributeError as e:
            logger.error(f"send_json - Attribute error: {e}")
        except ValueError as e:
            logger.error(f"send_json - Value error: {e}")
        except Exception as e:
            logger.error(f"send_json - unexpected error: {e}")

    # Save by openpyxl
    @tic_tok
    def save_to_excel(self) -> None:
        try:
            logger.info(f"save_to_excel - stock_number = {self._stock_code}")

            # 1) Preliminary
            today:datetime = datetime.today()
            root_path:str = "C:/temp/stock-log"
            filename:str = os.path.join(root_path, f"{self._stock_code}_{today.strftime('%Y-%m')}.xlsx")
            attrs:list[str] = ["created_at", "balance_yest", "selling_today", "return_today", "balance_today", "price"]

            wb = openpyxl.load_workbook(filename, data_only=True) if os.path.exists(filename) else Workbook()
            sheet:Worksheet = cast(Worksheet, wb.active)

            # 2) Check if data exist ?
            if os.path.exists(filename):
                logger.info(f"save_to_excel - file exists")
                # Create work book
                # wb:Workbook = openpyxl.load_workbook(filename, data_only=True)

                # Get first sheet when open the xlsx
                # sheet:Worksheet = cast(Worksheet, wb.active)

                # Get max_row number which is the index of latest record
                max_row:int = sheet.max_row
                if max_row != 1:
                    # Get last record datetime (string, yyyy-mm-dd)
                    # cast(type, val) => Tell mypy the type of val
                    lr_date:str = cast(str, sheet.cell(max_row, 1).value)[0:10]

                    # 3) Check if data duplicate ?
                    if lr_date == today.strftime("%Y-%m-%d"):  # compare with today
                        logger.info("save_to_excel - Duplicate date of record, stop saving data ")
                        return

            # 4) if file not exists
            else:
                logger.info(f"save_to_excel - file not exists, creating ...")

                # create new Excel file
                # sheet:Worksheet = cast(Worksheet, wb.active)
                sheet.append(attrs)
                wb.save(filename)

            # 5) build new row
            logger.info(f"save_to_excel - Building new row ...")
            row_arr:list[Union[str, float]] = []
            for attr in attrs:
                value:str = getattr(self, "_"+attr)
                if attr != "created_at" and value is not None:
                    row_arr.append(float(value.replace(",", "")))
                elif attr == "created_at":
                    row_arr.append(value)
                else:
                    row_arr.append("")

            sheet.append(row_arr)
            wb.save(filename)

            logger.info("save_to_excel - finish")

        except Exception as e:
            print(e)

    @tic_tok
    def send_chart(self) -> None:
        logger.info(f"send_chart - stock_number = {self._stock_code}")
        try:
            # 1) draw new chart
            plot_short_selling(self._stock_code)

            # 2) Path to the JPG file
            today:datetime = datetime.today()
            jpg_file_path:str = f"C:/temp/stock-log/{self._stock_code}_{today.strftime('%Y-%m')}.jpg"  # Replace with your actual file name

            # 3) Discord webhook URL
            url:str = "https://discord.com/api/webhooks/1356484738029719573/9GNCPHfl7gcz9BpkkO1xYEYqZ9_D2tWd0dx5sZqx3RTN3HgLFLql47TEgWYEsz0Q4x8g"

            # 4) Check if the file exists
            if not os.path.exists(jpg_file_path):
                logger.info(f"send_chart - File not found: {jpg_file_path}")
                return

            # 5) Prepare the file and payload
            with open(jpg_file_path, "rb") as file:
                files:dict = {"file": (os.path.basename(jpg_file_path), file, "image/jpeg")}
                payload:dict = {"username": "newmanBot"}

                # Send the request
                res = requests.post(url, data=payload, files=files)

            # 6) Read the response
            if res.status_code in (200, 204):
                logger.info(f"send_chart - success")
            else:
                logger.info(f"send_chart - fail")

        except Exception as e:
            logger.error(f"save_to_excel - error: {e}")

    def schedule_task(self) -> None:
        scheduler:BackgroundScheduler = BackgroundScheduler(timezone="Asia/Taipei")
        hour:int = 21
        minute:int = 30
        sec:int = 00

        logger.info(f"schedule_task init - stock_number = {self._stock_code}, time = {hour}:{minute}")

        scheduler.add_job(self.crawl_price, 'cron', day_of_week='mon-fri', hour=(hour - 1), minute=minute, second=sec)
        scheduler.add_job(self.crawl_short_selling, 'cron', day_of_week='mon-fri', hour=hour, minute=minute, second=sec)
        scheduler.add_job(self.save_to_excel, 'cron', day_of_week='mon-fri', hour=hour, minute=minute, second=(sec + 10) % 60)
        scheduler.add_job(self.send_json, 'cron', day_of_week='mon-fri', hour=hour, minute=minute, second=(sec + 20) % 60)
        scheduler.add_job(self.send_chart, 'cron', day_of_week='mon-fri', hour=hour, minute=minute, second=(sec + 30) % 60)
        scheduler.start()
